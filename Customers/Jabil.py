import tkinter as tk
from copy import copy
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter


def press_action():
    print("Hello partnership")


def jabil_logic():
    root = tk.Tk()
    root.title("Jabil Module")
    root.geometry("800x500")

    # Color scheme
    bg_color = '#3b5998'
    text_color = '#ffffff'
    button_color = '#8b9dc3'
    button_text_color = '#ffffff'

    root.configure(bg=bg_color)

    label = tk.Label(root, text="Welcome Partnership Member", font=("Verdana", 24), bg=bg_color, fg=text_color)
    label.pack(pady=20)

    instructions = ("Instructions:\n"
                    "1. Select the Award File for Jabil.\n"
                    "2. Select the BOM file for Jabil.\n"
                    "3. Save the final merged file.")
    instructions_label = tk.Label(root, text=instructions, font=("Verdana", 16), bg=bg_color, fg=text_color)
    instructions_label.pack(pady=20)

    process_btn = tk.Button(root, text="Process and Merge Files", command=lambda: process_and_merge_files(root),
                            bg=button_color, fg=button_text_color, font=("Verdana", 16))
    process_btn.pack(pady=10)

    root.mainloop()


def process_and_merge_files(parent_window):
    columns_to_merge = [
        'PSoft Part', 'Quoted Mfg', 'Quoted Part', 'Part Class', 'Last Ship Resale', 'Last Ship Date',
        'Last Ship GP', '12 Mo CPN Sales', 'Backlog Resale', 'Cust Backlog Value', 'Sager Stock', 'Stock Type',
        'On POs', 'Sager Min', 'Sager Mult', 'Factory LT', 'Avg Cost', 'Vol1 Cost', 'Vol2 Cost', 'Best Book',
        'Best Contract', 'Sager NCNR', 'Last PO Price', 'SND Cost', 'SND Quote', 'SND Exp Date', 'SND Cust ID',
        'VPC Cost', 'VPC Quote', 'VPC Exp Date', 'VPC MOQ', 'VPC TYPE', 'TIR MOQ', 'VPC Cust ID', 'Design Reg #',
        'Reg #', 'Last Ship CPN', 'Last Ship Cust ID', 'Backlog CPN', 'Backlog Entry', 'Backlog Cust ID'
    ]

    initial_dir = r"C:\Users\nabil\OneDrive\Documentos\WORKFILES\awardsprocess"

    file_path_1 = filedialog.askopenfilename(parent=parent_window, title="Select the first Excel file",
                                             filetypes=[("Excel files", "*.xlsx")], initialdir=initial_dir)
    file_path_2 = filedialog.askopenfilename(parent=parent_window, title="Select the second Excel file",
                                             filetypes=[("Excel files", "*.xlsm")], initialdir=initial_dir)
    if not file_path_1 or not file_path_2:
        messagebox.showwarning("Warning", "You need to select both files!")
        return

    try:
        # Process the first file
        workbook = load_workbook(file_path_1, data_only=False)
        sheet = workbook.active
        process_first_file(sheet)

        # Load the second file and add the 'Working Copy' sheet
        workbook2 = load_workbook(file_path_2, data_only=True)
        if 'Working Copy' in workbook2.sheetnames:
            sheet2 = workbook2['Working Copy']
            new_sheet = workbook.create_sheet('Working Copy')

            # Copying data and style from sheet2 to new_sheet
            for row in sheet2:
                for cell in row:
                    new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:  # Copy style if present
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)

            merge_columns(sheet, new_sheet, columns_to_merge)  # Call the merge function here

        else:
            messagebox.showwarning("Warning", "'Working Copy' sheet not found in the second file!")
            return

        # After all merging and processing are complete
        column_formats = {
            'AF': '"$"#,##0.00',  # Currency values with two decimal places
            'AI': '"$"#,##0.00',  # Currency values with two decimal places
            'AJ': '"$"#,##0.00',  # Currency values with two decimal places
            'CA': '"$"#,##0.00',  # Currency values with two decimal places
            'AK': '0.00%'  # Percentage format
        }
        format_columns(sheet, column_formats)

        # Ask user for save location and file name
        save_file_path = filedialog.asksaveasfilename(parent=parent_window, defaultextension=".xlsx",
                                                      filetypes=[("Excel files", "*.xlsx")],
                                                      title="Save Merged File As", initialdir=initial_dir)
        if save_file_path:
            workbook.save(save_file_path)
            messagebox.showinfo("Success", "File has been processed and saved successfully!")

    except Exception as e:
        messagebox.showerror("Error", "An error occurred during processing: " + str(e))


# This is the coloring function for this module that colors the rows, so it's similar to the original file
def fill_alternate_rows(sheet, start_row, end_row, color):
    columns_to_fill = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
                       'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE']

    for row in range(start_row, end_row + 1, 2):  # Start from start_row, increment by 2 until end_row
        for col_letter in columns_to_fill:
            cell = sheet[col_letter + str(row)]
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')


def process_first_file(sheet):
    # Initialize dictionary to map column headers to their respective columns
    sheet.insert_rows(1)
    header_column_map = {}
    for column in range(1, sheet.max_column + 1):
        header_value = sheet.cell(row=2, column=column).value  # Headers are in row 2, change the row value
        if header_value:
            header_column_map[header_value] = column

    # Print the header column map to see what's in there
    print("Header Column Map:", header_column_map)

    # Print out the headers in the first workbook
    print("Headers in the first workbook:", list(header_column_map.keys()))

    # Check for required headers and map them to column letters
    required_headers = ['Net Demand(180 Days)', 'Sourced Price', 'Sourced MOQ']
    missing_headers = [header for header in required_headers if header not in header_column_map]
    if missing_headers:
        messagebox.showerror("Error", f"Missing required columns: {', '.join(missing_headers)}")
        return None  # Exit if there are missing headers

    # Determine columns for new headers and formulas
    new_column_index = sheet.max_column + 1
    fill = PatternFill(start_color='FCD5B4', end_color='FCD5B4', fill_type='solid')
    wrap_text = Alignment(wrap_text=True)  # Enable text wrapping

    # List of new headers and their respective formula if applicable, here we created the new columns where we wil apply data and formulas
    new_headers = [
        ('Ext Award Value', None),
        ('Award Conf', None),
        ('180 Day Net', None),
        ('Award Price', None),
        ('Conf Cost', None),
        ('Award Margin', None),
        ('Award MOQ', None),
        ('Cost Comment', None),
        ('New Business', None)
    ]

    for i, (header, formula) in enumerate(new_headers):
        col_letter = get_column_letter(new_column_index + i)
        cell = sheet.cell(row=2, column=new_column_index + i)
        cell.value = header
        cell.fill = fill
        cell.alignment = wrap_text  # Apply wrap text alignment to each header cell
        if header == 'Conf Cost':
            for row in range(3, sheet.max_row + 1):  # Start processing from row 3 as row 2 contains headers
                sheet.cell(row=row, column=new_column_index + i).value = f'=BP{row}'  # Set Conf Cost = BY row value

    # Applying formulas
    try:
        awarded_eau_column = get_column_letter(
            header_column_map['Net Demand(180 Days)'])  # In other words its EAU for Jabil
        award_price_column = get_column_letter(header_column_map['Sourced Price'])
        moq_column = get_column_letter(header_column_map['Sourced MOQ'])
    except Exception as e:
        print("Error getting column letter:", e)
        return None

    print("Awarded EAU Column:", awarded_eau_column)
    print("Award Price Column:", award_price_column)
    print("MOQ Column:", moq_column)

    award_margin_column = get_column_letter(new_column_index + 6)  # Adjust the index for 'Award Margin'
    print("Award Margin Column:", award_margin_column)
    cost_comment_column_index = new_column_index + 7  # adjust this index based on actual position

    for row in range(3, sheet.max_row + 1):  # Start processing from row 3 as row 2 contains headers
        sheet.cell(row=row, column=new_column_index).value = f'={awarded_eau_column}{row}*{award_price_column}{row}'
        sheet.cell(row=row, column=new_column_index + 2).value = f'={awarded_eau_column}{row}'
        sheet.cell(row=row, column=new_column_index + 3).value = f'={award_price_column}{row}'
        # sheet.cell(row=row, column=new_column_index + 5).value = f'=BZ{row}*BX{row}'  # Ext Cost formula
        sheet.cell(row=row, column=new_column_index + 5).value = f'=(AI{row}-AJ{row})/AI{row}'  # Award Margin formula
        sheet.cell(row=row, column=new_column_index + 6).value = f'=R{row}'  # Award MOQ formula (Changed from AK to AL)
        sheet.cell(row=row, column=cost_comment_column_index).value = f'=BQ{row}'

    # Apply filters to the entire header row
    sheet.auto_filter.ref = f"A2:{get_column_letter(sheet.max_column)}2"

    fill_alternate_rows(sheet, start_row=4, end_row=sheet.max_row, color='ADD8E6')

    # Print to verify if all formulas and filters were applied correctly
    print("Auto filter applied to:", sheet.auto_filter.ref)

    # Add the SUBTOTAL formula in row 1 column AF cell with dynamic row reference
    max_row = sheet.max_row
    sheet.cell(row=1,
               column=new_column_index).value = f'=SUBTOTAL(9, AF3:AF{max_row})'  # Subtotal formula above Ext Value, we don't
    # add +1 here so this process happens right at the beginning of our column creation process

    # Print to verify if all formulas and filters were applied correctly
    print("Auto filter applied to:", sheet.auto_filter.ref)


def merge_columns(sheet1, working_copy, columns_to_merge):
    # Define the fill color and text wrapping for new column headers
    # fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    wrap_text = Alignment(wrap_text=True)

    # Identify the MPN column in both sheets
    mpn_column_index_sheet1 = None
    mpn_column_index_wc = None

    # Find MPN column in sheet1
    # Used Item number in sheet 1 and matched this to CPN in working copy to pull in data
    for col in range(1, sheet1.max_column + 1):
        if sheet1.cell(row=2, column=col).value == 'JPN':
            mpn_column_index_sheet1 = col
            break

    if not mpn_column_index_sheet1:
        messagebox.showerror("Error", "MPN column not found in the main sheet.")
        return

    # Find MPN column in the Working Copy
    for col in range(1, working_copy.max_column + 1):
        if working_copy.cell(row=3, column=col).value == 'CPN':
            mpn_column_index_wc = col
            break

    if not mpn_column_index_wc:
        messagebox.showerror("Error", "MPN column not found in the 'Working Copy'.")
        return

    # Mapping MPN values to row numbers in sheet1
    mpn_to_row = {}
    for row in range(3, sheet1.max_row + 1):
        mpn_value = sheet1.cell(row=row, column=mpn_column_index_sheet1).value
        mpn_to_row[mpn_value] = row

    # Mapping column headers to column indices in the Working Copy
    wc_column_indices = {}
    for col in range(1, working_copy.max_column + 1):
        header = working_copy.cell(row=3, column=col).value
        if header in columns_to_merge:
            wc_column_indices[header] = col

    # Adding new columns to sheet1 in row 2 and applying fill and text wrapping only to the headers
    next_column = sheet1.max_column + 1
    for header in columns_to_merge:
        cell = sheet1.cell(row=2, column=next_column)
        cell.value = header
        # cell.fill = fill
        cell.alignment = wrap_text
        next_column += 1

    # Filling the new columns with data from the 'Working Copy' based on MPN
    for row in range(4, working_copy.max_row + 1):  # Assuming data starts from row 4 in Working Copy
        wc_mpn_value = working_copy.cell(row=row, column=mpn_column_index_wc).value
        if wc_mpn_value in mpn_to_row:
            target_row = mpn_to_row[wc_mpn_value]
            for header, col_index in wc_column_indices.items():
                new_col = sheet1.max_column - len(columns_to_merge) + list(columns_to_merge).index(header) + 1
                sheet1.cell(row=target_row, column=new_col).value = working_copy.cell(row=row, column=col_index).value

    # Enable filters on the header row
    sheet1.auto_filter.ref = f"A2:{sheet1.cell(row=2, column=sheet1.max_column).coordinate}"


def format_columns(sheet, column_formats):
    for col_letter, format_style in column_formats.items():
        # Format the first row if the column is AX, BC, or BD
        if col_letter in ['AF', 'AI', 'AJ', 'AK', 'CA', 'CB', 'BZ']:  # Add 'BC' and 'BD' to the condition
            sheet.cell(row=1, column=openpyxl.utils.column_index_from_string(col_letter)).number_format = format_style

        # Format the rest of the column starting from row 3
        for row in range(3, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=openpyxl.utils.column_index_from_string(col_letter))
            cell.number_format = format_style
