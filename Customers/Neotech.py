import tkinter as tk
from copy import copy
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill


def press_action():
    print("Hello partnership")


def neotech_logic():
    root = tk.Tk()
    root.title("Neotech Module")
    root.geometry("800x500")

    # Color scheme
    bg_color = '#3b5998'
    text_color = '#ffffff'
    button_color = '#8b9dc3'
    button_text_color = '#ffffff'

    root.configure(bg=bg_color)

    label = tk.Label(root, text="Welcome Partnership Member", font=("Verdana", 24), bg=bg_color, fg=text_color)
    label.pack(pady=20)

    instructions = ("Instructions:\n"
                    "1. Select the Award File for Sanmina.\n"
                    "2. Select the BOM file for Sanmina.\n"
                    "3. Save the final merged file.")
    instructions_label = tk.Label(root, text=instructions, font=("Verdana", 16), bg=bg_color, fg=text_color)
    instructions_label.pack(pady=20)

    process_btn = tk.Button(root, text="Process and Merge Files", command=lambda: process_and_merge_files(root),
                            bg=button_color, fg=button_text_color, font=("Verdana", 16))
    process_btn.pack(pady=10)

    root.mainloop()


def process_and_merge_files(parent_window):
    columns_to_merge = [
        'PSoft Part', 'Quoted Mfg', 'Quoted Part', 'Part Class', 'Last Ship Resale', 'Last Ship Date',
        'Last Ship GP', '12 Mo CPN Sales', 'Backlog Resale', 'Cust Backlog Value', 'Sager Stock', 'Stock Type',
        'On POs', 'Sager Min', 'Sager Mult', 'Factory LT', 'Avg Cost', 'Vol1 Cost', 'Vol2 Cost', 'Best Book',
        'Best Contract', 'Sager NCNR', 'Last PO Price', 'SND Cost', 'SND Quote', 'SND Exp Date', 'SND Cust ID',
        'VPC Cost', 'VPC Quote', 'VPC Exp Date', 'VPC MOQ', 'VPC TYPE', 'TIR MOQ', 'VPC Cust ID', 'Design Reg #',
        'Reg #', 'Last Ship CPN', 'Last Ship Cust ID', 'Backlog CPN', 'Backlog Entry', 'Backlog Cust ID'
    ]

    initial_dir = r"C:\Users\nabil\OneDrive\Documentos\WORKFILES\awardsprocess"

    file_path_1 = filedialog.askopenfilename(parent=parent_window, title="Select the first Excel file",
                                             filetypes=[("Excel files", "*.xlsx")], initialdir=initial_dir)
    file_path_2 = filedialog.askopenfilename(parent=parent_window, title="Select the second Excel file",
                                             filetypes=[("Excel files", "*.xlsm")], initialdir=initial_dir)
    if not file_path_1 or not file_path_2:
        messagebox.showwarning("Warning", "You need to select both files!")
        return

    try:
        # Process the first file
        workbook = load_workbook(file_path_1, data_only=False)
        sheet = workbook.active
        process_first_file(sheet)

        # Load the second file and add the 'Working Copy' sheet
        workbook2 = load_workbook(file_path_2, data_only=True)
        if 'Working Copy' in workbook2.sheetnames:
            sheet2 = workbook2['Working Copy']
            new_sheet = workbook.create_sheet('Working Copy')

            # Copying data and style from sheet2 to new_sheet
            for row in sheet2:
                for cell in row:
                    new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:  # Copy style if present
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)

            merge_columns(sheet, new_sheet, columns_to_merge)  # Call the merge function here
            update_conf_cost(sheet)
            apply_conditional_formatting(sheet)  # Apply conditional formatting for Award Margin

        else:
            messagebox.showwarning("Warning", "'Working Copy' sheet not found in the second file!")
            return

        # After all merging and processing are complete
        column_formats = {
            'BV': '"$"#,##0.00',  # Currency values with two decimal places
            'BY': '"$"#,##0.00',  # Currency values with two decimal places
            'BZ': '"$"#,##0.00',  # Currency values with two decimal places
            'CA': '"$"#,##0.00',  # Currency values with two decimal places
            'CB': '0.00%'  # Percentage format
        }
        format_columns(sheet, column_formats)

        # Ask user for save location and file name
        save_file_path = filedialog.asksaveasfilename(parent=parent_window, defaultextension=".xlsx",
                                                      filetypes=[("Excel files", "*.xlsx")],
                                                      title="Save Merged File As", initialdir=initial_dir)
        if save_file_path:
            workbook.save(save_file_path)
            messagebox.showinfo("Success", "File has been processed and saved successfully!")

    except Exception as e:
        messagebox.showerror("Error", "An error occurred during processing: " + str(e))


def update_conf_cost(sheet):
    # Update the header map after any new columns are added
    header_column_map = {sheet.cell(row=2, column=col).value: col for col in range(1, sheet.max_column + 1)}

    # Check if 'Vol1 Cost' column is present
    if 'Vol1 Cost' not in header_column_map:
        print("Error: 'Vol1 Cost' column not found")
        return

    # Find index for 'Conf Cost' or create it if not exists
    if 'Conf Cost' not in header_column_map:
        conf_cost_col = sheet.max_column + 1
        sheet.cell(row=2, column=conf_cost_col).value = 'Conf Cost'
        sheet.cell(row=2, column=conf_cost_col).alignment = Alignment(wrap_text=True)
        header_column_map['Conf Cost'] = conf_cost_col
    else:
        conf_cost_col = header_column_map['Conf Cost']

    # 'Vol1 Cost' column index
    vol1_cost_col = header_column_map['Vol1 Cost']

    # Copy values from 'Vol1 Cost' to 'Conf Cost'
    for row in range(3, sheet.max_row + 1):
        vol1_cost_value = sheet.cell(row=row, column=vol1_cost_col).value
        sheet.cell(row=row, column=conf_cost_col).value = vol1_cost_value


def process_first_file(sheet):
    # Initialize dictionary to map column headers to their respective columns
    header_column_map = {sheet.cell(row=2, column=col).value: col
                         for col in range(1, sheet.max_column + 1) if sheet.cell(row=2, column=col).value}

    # Ensure required headers are present
    required_headers = ['Quoted Net Demand', 'Unit Price', 'MOQ']
    if any(header not in header_column_map for header in required_headers):
        missing = [header for header in required_headers if header not in header_column_map]
        print("Error: Missing required headers", missing)
        return

    # Define new headers and their formulas
    new_headers_formulas = {
        'Revised Resale': None,
        'Revised Margin': None,
        'Revised MOQ': None,
        'Flag for Increase': None,
        'Ext Award Value': f"={get_column_letter(header_column_map['Quoted Net Demand'])}{{row}}*{get_column_letter(header_column_map['Unit Price'])}{{row}}",
        'Award Conf': None,
        'EAU': f"={get_column_letter(header_column_map['Quoted Net Demand'])}{{row}}",
        'Award Price': f"={get_column_letter(header_column_map['Unit Price'])}{{row}}",
        'Conf Cost': None,  # Assuming this will be populated later
        'Ext Cost': None,  # To be calculated after mapping update
        'Award Margin': None,  # To be calculated with updated formula below
        'Award MOQ': f"={get_column_letter(header_column_map['MOQ'])}{{row}}",  # Copy from 'MOQ',
        'Cost Comment': None,
        'New Business': None
    }

    # Start inserting new columns from the next available column
    start_column = sheet.max_column + 1

    # Insert headers and apply formulas where applicable
    for header, formula in new_headers_formulas.items():
        col_letter = get_column_letter(start_column)
        sheet.cell(row=2, column=start_column).value = header

        # Set fill color based on the header name
        if header in ['Revised Resale', 'Revised Margin', 'Revised MOQ', 'Flag for Increase']:
            fill_color = 'FFFFFF'  # White
        else:
            fill_color = 'FCD5B4'  # The previous color used for other headers

        sheet.cell(row=2, column=start_column).fill = PatternFill(start_color=fill_color, fill_type='solid')
        sheet.cell(row=2, column=start_column).alignment = Alignment(wrap_text=True)

        # Update the header column map immediately after adding each header
        header_column_map[header] = start_column

        if formula:  # Apply formulas to all rows starting from row 3
            for row in range(3, sheet.max_row + 1):
                sheet.cell(row=row, column=start_column).value = formula.format(row=row)

        start_column += 1  # Increment column for next header

    # Calculate 'Ext Cost' and 'Award Margin'
    conf_cost_column = header_column_map['Conf Cost']
    eau_column = header_column_map['EAU']
    award_price_column = header_column_map['Award Price']
    ext_cost_column = header_column_map['Ext Cost']
    award_margin_column = header_column_map['Award Margin']

    for row in range(3, sheet.max_row + 1):
        # Calculate Ext Cost if Conf Cost and EAU columns are populated
        if conf_cost_column and eau_column:
            sheet.cell(row=row,
                       column=ext_cost_column).value = f"={get_column_letter(conf_cost_column)}{row}*{get_column_letter(eau_column)}{row}"

        # Calculate Award Margin if Award Price and Conf Cost are populated
        if award_price_column and conf_cost_column:
            sheet.cell(row=row, column=award_margin_column).value = \
                f"=({get_column_letter(award_price_column)}{row}-{get_column_letter(conf_cost_column)}{row})/{get_column_letter(award_price_column)}{row}"

    # Apply subtotal formula to the 'Ext Award Value' column
    ext_award_value_col = header_column_map['Ext Award Value']
    last_row = sheet.max_row
    ext_cost_col_letter = get_column_letter(ext_cost_column)
    subtotal_formula = f"=SUBTOTAL(9, {get_column_letter(ext_award_value_col)}3:{get_column_letter(ext_award_value_col)}{last_row})"
    sheet.cell(row=1, column=ext_award_value_col).value = subtotal_formula  # Placing the formula in the first row

    # Subtotal formula for 'Ext Cost' above its header
    subtotal_formula_ext_cost = f"=SUBTOTAL(9, {ext_cost_col_letter}3:{ext_cost_col_letter}{last_row})"
    sheet.cell(row=1,
               column=ext_cost_column).value = subtotal_formula_ext_cost  # Placing the formula above the column header

    # Assume 'BV' and 'CA' are your actual Excel column headers or replace them accordingly
    bv_column_letter = get_column_letter(header_column_map['Ext Award Value'])  # Replace with your actual column header
    ca_column_letter = get_column_letter(header_column_map['Ext Cost'])
    award_margin_column = header_column_map['Award Margin']

    # Correct formula setting
    award_margin_formula = f"=({bv_column_letter}1-{ca_column_letter}1)/{bv_column_letter}1"
    sheet.cell(row=1, column=award_margin_column).value = award_margin_formula

    # Apply filters to the header row
    sheet.auto_filter.ref = f"A2:{get_column_letter(sheet.max_column)}2"


def merge_columns(sheet1, working_copy, columns_to_merge):
    # Define the fill color and text wrapping for new column headers
    wrap_text = Alignment(wrap_text=True)

    # Identify the MPN column in both sheets
    mpn_column_index_sheet1 = None
    mpn_column_index_wc = None

    # Find MPN column in sheet1
    for col in range(1, sheet1.max_column + 1):
        if sheet1.cell(row=2, column=col).value == 'PartNum':
            mpn_column_index_sheet1 = col
            break

    if not mpn_column_index_sheet1:
        messagebox.showerror("Error", "PartNum column not found in the main sheet.")
        return

    # Find MPN column in the Working Copy
    for col in range(1, working_copy.max_column + 1):
        if working_copy.cell(row=3, column=col).value == 'CPN':
            mpn_column_index_wc = col
            break

    if not mpn_column_index_wc:
        messagebox.showerror("Error", "CPN column not found in the 'Working Copy'.")
        return

    # Mapping MPN values to row numbers in sheet1
    mpn_to_row = {}
    for row in range(3, sheet1.max_row + 1):
        mpn_value = sheet1.cell(row=row, column=mpn_column_index_sheet1).value
        mpn_to_row[mpn_value] = row

    # Mapping column headers to column indices in the Working Copy
    wc_column_indices = {}
    for col in range(1, working_copy.max_column + 1):
        header = working_copy.cell(row=3, column=col).value
        if header in columns_to_merge:
            wc_column_indices[header] = col

        # Add new columns to sheet1 and fill them with data from the 'Working Copy'
    next_column = sheet1.max_column + 1
    psoft_part_index = None
    for header in columns_to_merge:
        cell = sheet1.cell(row=2, column=next_column)
        cell.value = header
        cell.alignment = wrap_text
        if header == "PSoft Part":
            psoft_part_index = next_column
        next_column += 1

    for row in range(4, working_copy.max_row + 1):
        wc_mpn_value = working_copy.cell(row=row, column=mpn_column_index_wc).value
        if wc_mpn_value in mpn_to_row:
            target_row = mpn_to_row[wc_mpn_value]
            for header, col_index in wc_column_indices.items():
                new_col = sheet1.max_column - len(columns_to_merge) + list(columns_to_merge).index(header) + 1
                sheet1.cell(row=target_row, column=new_col).value = working_copy.cell(row=row, column=col_index).value

    # Add 'PSID Ct' column next to 'PSoft Part' and apply COUNTIF formula
    if psoft_part_index:
        psid_ct_col = psoft_part_index + 1
        sheet1.cell(row=2, column=psid_ct_col).value = 'PSID Ct'
        sheet1.cell(row=2, column=psid_ct_col).alignment = wrap_text
        psoft_part_col_letter = get_column_letter(psoft_part_index)
        for row in range(3, sheet1.max_row + 1):
            countif_formula = f"=COUNTIF({psoft_part_col_letter}:{psoft_part_col_letter}, {psoft_part_col_letter}{row})"
            sheet1.cell(row=row, column=psid_ct_col).value = countif_formula

    # Enable filters on the header row
    sheet1.auto_filter.ref = f"A2:{get_column_letter(sheet1.max_column)}2"


def format_columns(sheet, column_formats):
    for col_letter, format_style in column_formats.items():
        # Format the first row if the column is AX, BC, or BD
        if col_letter in ['AX', 'BC', 'BD', 'BV', 'CA', 'CB', 'BZ']:  # Add 'BC' and 'BD' to the condition
            sheet.cell(row=1, column=openpyxl.utils.column_index_from_string(col_letter)).number_format = format_style

        # Format the rest of the column starting from row 3
        for row in range(3, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=openpyxl.utils.column_index_from_string(col_letter))
            cell.number_format = format_style


def apply_conditional_formatting(sheet):
    # Find the column indices for 'Award Margin' and 'Conf Cost'
    award_margin_col = None
    conf_cost_col = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=2, column=col).value == 'Award Margin':
            award_margin_col = col
        elif sheet.cell(row=2, column=col).value == 'Conf Cost':
            conf_cost_col = col

    if award_margin_col:
        # Apply conditional formatting to highlight Award Margin values below 6% with light red
        light_red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        rule = CellIsRule(operator='lessThan', formula=['0.06'], stopIfTrue=True, fill=light_red_fill)
        sheet.conditional_formatting.add(
            f"{get_column_letter(award_margin_col)}3:{get_column_letter(award_margin_col)}{sheet.max_row}", rule)
